import os
import re
import json
from io import BytesIO
from datetime import datetime

from flask import Flask, request, send_file, jsonify, make_response

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import FormulaRule


# ============================================================
# Flask app
# ============================================================
app = Flask(__name__)

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def _cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS, GET"
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["X-AUREA"] = "excel"
    return resp


def _safe_filename(name: str) -> str:
    name = (name or "AUREA_excel.xlsx").strip()
    name = re.sub(r"[^\w.\- ]+", "_", name)
    if not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name


def _to_dict(x):
    """
    Normaliza cualquier entrada rara a dict para evitar:
    'list' object has no attribute 'setdefault'
    """
    if x is None:
        return {}
    if isinstance(x, dict):
        return x
    if isinstance(x, str):
        try:
            v = json.loads(x)
            return v if isinstance(v, dict) else {"value": v}
        except Exception:
            return {"value": x}
    if isinstance(x, list):
        return {"items": x}
    return {"value": x}


def _as_number(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace("MXN", "").replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def _looks_like_services_template(text: str) -> bool:
    if not text:
        return False
    t = text.lower()
    keys = ["precio fijo", "estudios realizados", "total generado", "servicios", "tomografia", "ultrasonido", "rayos x"]
    score = sum(1 for k in keys if k in t)
    return score >= 2


def _now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ------------------------------------------------------------
# Styling helpers
# ------------------------------------------------------------
NAVY = "0B1220"
NAVY_2 = "111B2E"
GOLD = "D6B25E"
WHITE = "FFFFFF"

thin = Side(style="thin", color="263043")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

money_fmt = '"MXN" #,##0.00'

def style_title(cell):
    cell.font = Font(bold=True, size=14, color=GOLD)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_header_row(ws, row, start_col, end_col):
    fill = PatternFill("solid", fgColor=NAVY_2)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ============================================================
# Excel builder (ledger + dashboard + charts + alerts)
# ============================================================
def build_excel(payload: dict) -> BytesIO:
    payload = _to_dict(payload)

    prompt = str(payload.get("prompt") or payload.get("text") or "")
    file_name = _safe_filename(payload.get("fileName") or payload.get("filename") or "AUREA_excel.xlsx")

    rows = payload.get("rows")
    rows = rows if isinstance(rows, list) else []

    use_services_template = _looks_like_services_template(prompt)

    wb = Workbook()
    ws = wb.active
    ws.title = "AUREA"
    dash = wb.create_sheet("Dashboard")

    # hoja listas oculta (validaciones pro)
    lists = wb.create_sheet("_lists")
    lists.sheet_state = "hidden"

    # look
    ws.sheet_view.showGridLines = True
    dash.sheet_view.showGridLines = True

    # Freeze panes
    ws.freeze_panes = "A4"
    dash.freeze_panes = "A4"

    # Header
    ws["A1"] = "AUREA 33 • AUREA"
    style_title(ws["A1"])
    ws.merge_cells("A1:F1")

    ws["A2"] = "Generado:"
    ws["B2"] = _now_str()
    ws["A2"].font = Font(bold=True, color=WHITE)
    ws["B2"].font = Font(color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["B2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    # ============================================================
    # TEMPLATE 1: Ledger contable PRO
    # ============================================================
    if not use_services_template:
        headers = ["Fecha", "Concepto", "Categoría", "Forma de pago", "Ingreso", "Egreso"]
        header_row = 3
        for i, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=i, value=h)
        style_header_row(ws, header_row, 1, 6)

        # rango data
        data_first = header_row + 1
        data_rows = 5000
        data_last = data_first + data_rows - 1
        total_row = data_last + 1  # fila de totales dentro de la tabla

        # columnas
        set_col_widths(ws, {
            "A": 14, "B": 24, "C": 16, "D": 18, "E": 14, "F": 14
        })

        # -------------------------
        # Listas para validación
        # -------------------------
        payments = ["Efectivo", "Transferencia", "Depósito", "Tarjeta Débito", "Tarjeta Crédito"]
        categorias_default = ["Alimentos", "Servicios", "Transporte", "Salud", "Hogar", "Negocio", "Educación", "Ocio", "Otros"]

        lists["A1"] = "Pagos"
        lists["B1"] = "Categorias"
        lists["A1"].font = Font(bold=True)
        lists["B1"].font = Font(bold=True)

        for i, p in enumerate(payments, start=2):
            lists[f"A{i}"] = p
        for i, c in enumerate(categorias_default, start=2):
            lists[f"B{i}"] = c

        pay_range = f"_lists!$A$2:$A${len(payments)+1}"
        cat_range = f"_lists!$B$2:$B${len(categorias_default)+1}"

        # Validación pagos (col D)
        dv_pay = DataValidation(type="list", formula1=f"={pay_range}", allow_blank=True)
        dv_pay.prompt = "Selecciona forma de pago"
        dv_pay.error = "Elige un valor del listado."
        ws.add_data_validation(dv_pay)
        dv_pay.add(f"D{data_first}:D{data_last}")

        # Validación categorías (col C)
        dv_cat = DataValidation(type="list", formula1=f"={cat_range}", allow_blank=True)
        dv_cat.prompt = "Selecciona categoría"
        dv_cat.error = "Elige una categoría del listado."
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"C{data_first}:C{data_last}")

        # -------------------------
        # Formatos / bordes base
        # -------------------------
        for r in range(data_first, data_last + 1):
            for c in range(1, 7):
                cell = ws.cell(row=r, column=c)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=r, column=5).number_format = money_fmt
            ws.cell(row=r, column=6).number_format = money_fmt

        # -------------------------
        # Prefill rows (opcional)
        # -------------------------
        write_r = data_first
        for item in rows[:300]:
            if not isinstance(item, dict):
                continue
            ws.cell(write_r, 1, item.get("Fecha") or item.get("fecha"))
            ws.cell(write_r, 2, item.get("Concepto") or item.get("concepto"))
            ws.cell(write_r, 3, item.get("Categoría") or item.get("categoria") or item.get("category"))
            ws.cell(write_r, 4, item.get("Forma de pago") or item.get("pago") or item.get("payment"))
            ws.cell(write_r, 5, _as_number(item.get("Ingreso") or item.get("ingreso")))
            ws.cell(write_r, 6, _as_number(item.get("Egreso") or item.get("egreso")))
            write_r += 1

        # -------------------------
        # Excel Table REAL (tbl_data)
        # Incluye Totals Row (para sumar) pero KPIs usan #Data => NO DUPLICA
        # -------------------------
        table_ref = f"A{header_row}:F{total_row}"
        tab = Table(displayName="tbl_data", ref=table_ref)

        # columnas con total function
        tab.tableColumns = [
            TableColumn(id=1, name="Fecha"),
            TableColumn(id=2, name="Concepto"),
            TableColumn(id=3, name="Categoría"),
            TableColumn(id=4, name="Forma de pago"),
            TableColumn(id=5, name="Ingreso", totalsRowFunction="sum"),
            TableColumn(id=6, name="Egreso", totalsRowFunction="sum"),
        ]
        tab.totalsRowCount = 1

        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tab)

        # estilo de la totals row (fila total_row)
        for c in range(1, 7):
            cell = ws.cell(row=total_row, column=c)
            cell.fill = PatternFill("solid", fgColor=NAVY_2)
            cell.font = Font(bold=True, color=WHITE)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=total_row, column=5).number_format = money_fmt
        ws.cell(row=total_row, column=6).number_format = money_fmt

        # -------------------------
        # Reglas de calidad (condicional)
        # 1) Falta categoría si hay monto
        # 2) Ingreso y Egreso a la vez (error humano)
        # -------------------------
        # Falta categoría si (Ingreso>0 o Egreso>0) y C vacío
        ws.conditional_formatting.add(
            f"C{data_first}:C{data_last}",
            FormulaRule(
                formula=[f'=AND($C{data_first}="",OR($E{data_first}>0,$F{data_first}>0))'],
                fill=PatternFill("solid", fgColor="7A1E1E")
            )
        )
        # Ingreso y Egreso en la misma fila
        ws.conditional_formatting.add(
            f"E{data_first}:F{data_last}",
            FormulaRule(
                formula=[f"=AND($E{data_first}>0,$F{data_first}>0)"],
                fill=PatternFill("solid", fgColor="7A1E1E")
            )
        )

        # ============================================================
        # Dashboard PRO (Structured Refs: NO DUPLICA JAMÁS)
        # ============================================================
        dash["A1"] = "AUREA 33 • Dashboard"
        style_title(dash["A1"])
        dash.merge_cells("A1:F1")

        # KPI header
        dash["A3"] = "KPI"
        dash["B3"] = "Valor"
        style_header_row(dash, 3, 1, 2)

        dash["A4"] = "Ingresos"
        dash["A5"] = "Egresos"
        dash["A6"] = "Balance"
        dash["A7"] = "Ingresos (Efectivo)"
        dash["A8"] = "Egresos (Tarjeta)"

        # ✅ Fórmulas con #Data (solo datos, excluye totals row)
        dash["B4"] = '=SUM(tbl_data[[#Data],[Ingreso]])'
        dash["B5"] = '=SUM(tbl_data[[#Data],[Egreso]])'
        dash["B6"] = "=B4-B5"
        dash["B7"] = '=SUMIFS(tbl_data[[#Data],[Ingreso]],tbl_data[[#Data],[Forma de pago]],"Efectivo")'
        dash["B8"] = '=SUMIFS(tbl_data[[#Data],[Egreso]],tbl_data[[#Data],[Forma de pago]],"Tarjeta*")'

        for r in range(4, 9):
            dash[f"A{r}"].border = border_thin
            dash[f"B{r}"].border = border_thin
            dash[f"A{r}"].fill = PatternFill("solid", fgColor=NAVY)
            dash[f"B{r}"].fill = PatternFill("solid", fgColor=NAVY)
            dash[f"A{r}"].font = Font(color=WHITE, bold=True)
            dash[f"B{r}"].font = Font(color=WHITE, bold=True)
            dash[f"B{r}"].number_format = money_fmt
            dash[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
            dash[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")

        dash.column_dimensions["A"].width = 24
        dash.column_dimensions["B"].width = 16

        # Alertas
        dash["D3"] = "Alertas"
        dash.merge_cells("D3:F3")
        dash["D3"].fill = PatternFill("solid", fgColor=NAVY_2)
        dash["D3"].font = Font(color=WHITE, bold=True)
        dash["D3"].alignment = Alignment(horizontal="center", vertical="center")

        dash["D4"] = "Balance negativo"
        dash["D5"] = "Egresos >= 5000"
        dash["D6"] = "Sin categoría"

        dash["E4"] = "=IF(B6<0,1,0)"
        dash["E5"] = '=COUNTIF(tbl_data[[#Data],[Egreso]],">=5000")'
        dash["E6"] = '=COUNTIF(tbl_data[[#Data],[Categoría]],"")'

        dash["F4"] = '=IF(E4=1,"⚠️","✅")'
        dash["F5"] = '=IF(E5>0,"⚠️","✅")'
        dash["F6"] = '=IF(E6>0,"⚠️ Completar","✅")'

        for r in range(4, 7):
            for col in ["D", "E", "F"]:
                cell = dash[f"{col}{r}"]
                cell.border = border_thin
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color=WHITE, bold=(col != "E"))
                cell.alignment = Alignment(horizontal="center" if col != "D" else "left", vertical="center")

        dash.column_dimensions["D"].width = 18
        dash.column_dimensions["E"].width = 12
        dash.column_dimensions["F"].width = 14

        dash.conditional_formatting.add(
            "B6",
            FormulaRule(formula=["B6<0"], fill=PatternFill("solid", fgColor="7A1E1E"))
        )

        # Gastos por categoría + Pie
        dash["D9"] = "Gastos por Categoría"
        dash.merge_cells("D9:E9")
        dash["D9"].fill = PatternFill("solid", fgColor=NAVY_2)
        dash["D9"].font = Font(color=WHITE, bold=True)
        dash["D9"].alignment = Alignment(horizontal="center", vertical="center")

        dash["D10"] = "Categoría"
        dash["E10"] = "Total"
        style_header_row(dash, 10, 4, 5)

        base_row = 11
        for i, cat in enumerate(categorias_default):
            r = base_row + i
            dash[f"D{r}"] = cat
            dash[f"E{r}"] = f'=SUMIFS(tbl_data[[#Data],[Egreso]],tbl_data[[#Data],[Categoría]],"{cat}")'
            dash[f"D{r}"].border = border_thin
            dash[f"E{r}"].border = border_thin
            dash[f"D{r}"].fill = PatternFill("solid", fgColor=NAVY)
            dash[f"E{r}"].fill = PatternFill("solid", fgColor=NAVY)
            dash[f"D{r}"].font = Font(color=WHITE, bold=True)
            dash[f"E{r}"].font = Font(color=WHITE, bold=True)
            dash[f"E{r}"].number_format = money_fmt

        dash.column_dimensions["D"].width = 16
        dash.column_dimensions["E"].width = 14

        pie = PieChart()
        pie.title = "Distribución de Egresos"
        labels = Reference(dash, min_col=4, min_row=base_row, max_row=base_row + len(categorias_default) - 1)
        data_ref = Reference(dash, min_col=5, min_row=base_row - 1, max_row=base_row + len(categorias_default) - 1)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 10
        pie.width = 20
        pie.legend.position = "r"
        pie.dataLabels = None
        dash.add_chart(pie, "D21")

        # Top 10 Egresos
        dash["A10"] = "Top 10 Egresos"
        dash.merge_cells("A10:C10")
        dash["A10"].fill = PatternFill("solid", fgColor=NAVY_2)
        dash["A10"].font = Font(color=WHITE, bold=True)
        dash["A10"].alignment = Alignment(horizontal="center", vertical="center")

        dash["A11"] = "Concepto"
        dash["B11"] = "Categoría"
        dash["C11"] = "Egreso"
        style_header_row(dash, 11, 1, 3)

        top_start = 12
        for i in range(10):
            r = top_start + i
            dash[f"C{r}"] = f"=LARGE(tbl_data[[#Data],[Egreso]],{i+1})"
            dash[f"A{r}"] = f'=IFERROR(INDEX(tbl_data[[#Data],[Concepto]],MATCH(C{r},tbl_data[[#Data],[Egreso]],0)),"")'
            dash[f"B{r}"] = f'=IFERROR(INDEX(tbl_data[[#Data],[Categoría]],MATCH(C{r},tbl_data[[#Data],[Egreso]],0)),"")'
            for col in ["A", "B", "C"]:
                cell = dash[f"{col}{r}"]
                cell.border = border_thin
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color=WHITE, bold=(col == "C"))
                cell.alignment = Alignment(horizontal="left", vertical="center")
            dash[f"C{r}"].number_format = money_fmt

        dash.column_dimensions["A"].width = 24
        dash.column_dimensions["B"].width = 16
        dash.column_dimensions["C"].width = 14

        bar = BarChart()
        bar.title = "Top Egresos"
        bar.height = 8
        bar.width = 18
        cats = Reference(dash, min_col=1, min_row=top_start, max_row=top_start + 9)
        vals = Reference(dash, min_col=3, min_row=11, max_row=top_start + 9)
        bar.add_data(vals, titles_from_data=True)
        bar.set_categories(cats)
        bar.legend = None
        dash.add_chart(bar, "A23")

    # ============================================================
    # TEMPLATE 2: Servicios (se mantiene, solo mejorado leve)
    # ============================================================
    else:
        ws["A3"] = "Plantilla de Servicios (precio fijo × estudios realizados)"
        ws["A3"].font = Font(bold=True, color=WHITE)
        ws["A3"].fill = PatternFill("solid", fgColor=NAVY)
        ws.merge_cells("A3:F3")

        headers = ["Servicio", "Nombre del servicio", "Precio unitario", "Estudios realizados", "Total generado", "Notas"]
        start_row = 5
        for i, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=i, value=h)
        style_header_row(ws, start_row, 1, 6)

        data_first = start_row + 1
        data_last = data_first + 299

        set_col_widths(ws, {
            "A": 22, "B": 30, "C": 16, "D": 18, "E": 16, "F": 20
        })

        servicios_default = [
            "Tomografía simple",
            "Tomografía contrastada",
            "Angiotomografía",
            "Ultrasonido básico",
            "Doppler / Ecografía Doppler",
            "Rayos X digital",
        ]

        # listas en _lists
        lists["D1"] = "Servicios"
        lists["D1"].font = Font(bold=True)
        for i, s in enumerate(servicios_default, start=2):
            lists[f"D{i}"] = s
        serv_range = f"_lists!$D$2:$D${len(servicios_default)+1}"

        dv_serv = DataValidation(type="list", formula1=f"={serv_range}", allow_blank=True)
        ws.add_data_validation(dv_serv)
        dv_serv.add(f"A{data_first}:A{data_last}")

        for r in range(data_first, data_last + 1):
            ws.cell(r, 5).value = f"=C{r}*D{r}"
            for c in range(1, 7):
                ws.cell(r, c).border = border_thin
                ws.cell(r, c).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(r, 3).number_format = money_fmt
            ws.cell(r, 5).number_format = money_fmt

        dash["A1"] = "AUREA 33 • Dashboard (Servicios)"
        style_title(dash["A1"])
        dash.merge_cells("A1:E1")

        dash["A3"] = "Total generado"
        dash["B3"] = f"=SUM(AUREA!E{data_first}:AUREA!E{data_last})"
        dash["B3"].number_format = money_fmt
        dash["A3"].font = Font(bold=True)
        dash["B3"].font = Font(bold=True)

        dash["A5"] = "Servicio"
        dash["B5"] = "Total"
        style_header_row(dash, 5, 1, 2)

        for i in range(len(servicios_default)):
            r = 6 + i
            dash[f"A{r}"] = servicios_default[i]
            dash[f"B{r}"] = f'=SUMIF(AUREA!A{data_first}:AUREA!A{data_last},"{servicios_default[i]}",AUREA!E{data_first}:AUREA!E{data_last})'
            dash[f"B{r}"].number_format = money_fmt
            dash[f"A{r}"].border = border_thin
            dash[f"B{r}"].border = border_thin

        bar = BarChart()
        bar.title = "Total por servicio"
        bar.height = 10
        bar.width = 20
        cats = Reference(dash, min_col=1, min_row=6, max_row=6 + len(servicios_default) - 1)
        vals = Reference(dash, min_col=2, min_row=5, max_row=6 + len(servicios_default) - 1)
        bar.add_data(vals, titles_from_data=True)
        bar.set_categories(cats)
        bar.legend = None
        dash.add_chart(bar, "A12")

    # Save
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    bio.name = file_name
    return bio


# ============================================================
# Routes
# ============================================================
@app.route("/healthz", methods=["GET"])
def healthz():
    return _cors(jsonify({"ok": True, "service": "aurea-excel-generator"}))


@app.route("/api/excel/generate", methods=["POST", "OPTIONS"])
def generate_excel():
    if request.method == "OPTIONS":
        return _cors(make_response("", 204))

    try:
        payload = request.get_json(silent=True) or {}
        payload = _to_dict(payload)

        out = build_excel(payload)
        filename = _safe_filename(payload.get("fileName") or "AUREA_excel.xlsx")

        resp = send_file(
            out,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return _cors(resp)

    except Exception as e:
        return _cors(jsonify({
            "ok": False,
            "error": f"excel_build_failed: {str(e)}"
        })), 500


# ============================================================
# Local run
# ============================================================
if __name__ == "__main__":
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8080"))
    debug = os.getenv("DEBUG", "1") == "1"
    app.run(host=host, port=port, debug=debug)
